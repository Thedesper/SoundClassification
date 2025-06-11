from mcp.server.fastmcp import FastMCP
import os
import matplotlib.pyplot as plt
import subprocess
from openpyxl import load_workbook
import OLGAOutput2xlsx

# Create MCP Server
mcp = FastMCP("Olga_Operator")

# Global default paths
DEFAULT_MODEL_PATH = r"C:\Users\Engineer\Desktop\OlgaMCPServer\model"
DEFAULT_DIAMETER = 0.1
nthreads = 4
olga_bin = r"C:\Program Files\Schlumberger\Olga 2025.1.0\OlgaExecutables\Olga-2025.1.0"
LICENSE_PATH = r"C:\Users\Engineer\Desktop\BXu4_08002742A532_20250528112511_001.lic"

### Utility Functions ###

@mcp.tool()
def update_diameter(diameter: float = DEFAULT_DIAMETER, pipeline_path: str = DEFAULT_MODEL_PATH) -> dict:
    """
    Update the pipeline diameter parameter in the Genkey file
    Args:
        diameter: Pipeline diameter (unit: meter)
        pipeline_path: Directory path of the model files
    Returns:
        dict: Operation result
    """
    result = {"status": "success", "message": ""}
    try:
        genkey_file_path = os.path.join(pipeline_path, "Basic2.genkey")
        if not os.path.exists(genkey_file_path):
            result["status"] = "error"
            result["message"] = f"Genkey file does not exist: {genkey_file_path}"
            return result
        
        with open(genkey_file_path, "r", encoding="utf-8") as file:
            content = file.read()
        
        if "DIAMETER=" in content:
            updated_content = content.replace(
                f"DIAMETER={content.split('DIAMETER=')[1].splitlines()[0]}",
                f"DIAMETER={diameter} m"
            )
            with open(genkey_file_path, "w", encoding="utf-8") as file:
                file.write(updated_content)
            result["message"] = f"Successfully updated diameter to {diameter} m"
            return result
        else:
            result["status"] = "error"
            result["message"] = "DIAMETER parameter not found in Genkey file"
            return result
    except Exception as e:
        result["status"] = "error"
        result["message"] = f"Failed to update diameter: {str(e)}"
        return result

@mcp.tool()
def run_olga_simulation(pipeline_path: str = DEFAULT_MODEL_PATH) -> dict:
    """
    Run OLGA simulation program
    Args:
        pipeline_path: Directory path of the model files
    Returns:
        dict: Operation result
    """
    result = {"status": "success", "message": ""}
    try:
        env = os.environ.copy()
        env["LM_LICENSE_FILE"] = LICENSE_PATH
        
        executed_files = []
        for subdir, _, files in os.walk(pipeline_path):
            for file in files:
                if file.endswith('.key') or file.endswith('.genkey'):
                    full_path = os.path.join(subdir, file)
                    cmd = f'"{olga_bin}" -nthreads {nthreads} "{full_path}"'
                    
                    # Execute command and capture output
                    process = subprocess.run(
                        cmd,
                        env=env,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        check=True
                    )
                    executed_files.append(full_path)
        
        if not executed_files:
            result["status"] = "warning"
            result["message"] = "No .key or .genkey files found"
        else:
            result["message"] = f"Successfully ran {len(executed_files)} simulation files"
        
        return result
    except subprocess.CalledProcessError as e:
        result["status"] = "error"
        result["message"] = f"Simulation execution failed: {e.stderr}"
        return result
    except Exception as e:
        result["status"] = "error"
        result["message"] = f"An unexpected error occurred while running simulation: {str(e)}"
        return result

@mcp.tool()
def process_olga_results(pipeline_path: str = DEFAULT_MODEL_PATH) -> dict:
    """
    Process OLGA output results and convert to Excel
    Args:
        pipeline_path: Directory path of the model files
    Returns:
        dict: Operation result
    """
    result = {"status": "success", "message": ""}
    try:
        # Call OLGAOutput2xlsx to convert results
        path_ = [pipeline_path]
        OLGAOutput2xlsx.testtpl(path_)
        
        excel_path = os.path.join(pipeline_path, "Basic2.tpl.xlsx")
        if os.path.exists(excel_path):
            result["message"] = f"Results converted to Excel: {excel_path}"
            result["excel_path"] = excel_path
            return result
        else:
            result["status"] = "error"
            result["message"] = "Excel result file not generated"
            return result
    except Exception as e:
        result["status"] = "error"
        result["message"] = f"Failed to process results: {str(e)}"
        return result

@mcp.tool()
def generate_qgst_plot(diameter: float = DEFAULT_DIAMETER, excel_path: str = None) -> dict:
    """
    Generate QGST chart from Excel results
    Args:
        diameter: Pipeline diameter (for chart label)
        excel_path: Input Excel file path (default to file in default model path)
    Returns:
        dict: Operation result
    """
    result = {"status": "success", "message": ""}
    if excel_path is None:
        excel_path = os.path.join(DEFAULT_MODEL_PATH, "Basic2.tpl.xlsx")
    
    try:
        if not os.path.exists(excel_path):
            result["status"] = "error"
            result["message"] = f"Excel file does not exist: {excel_path}"
            return result
        
        # Constant definitions
        SHEET_NAME = 'Data'
        TIME_COL = 2     # Column for time data (Column B)
        VALUE_COL = 23   # Column for value data (Column V)
        START_ROW = 8
        
        # Load workbook and worksheet
        wb = load_workbook(filename=excel_path)
        ws = wb[SHEET_NAME]
        
        # Read data
        time_data = []
        value_data = []
        
        for col in ws.iter_cols(min_row=START_ROW, min_col=TIME_COL, max_col=TIME_COL, values_only=True):
            time_data.extend([cell for cell in col if cell is not None])
        
        for col2 in ws.iter_cols(min_row=START_ROW, min_col=VALUE_COL, max_col=VALUE_COL, values_only=True):
            value_data.extend([cell for cell in col2 if cell is not None])
        
        # Ensure time data and value data have the same length
        min_len = min(len(time_data), len(value_data))
        time_data = time_data[:min_len]
        value_data = value_data[:min_len]
        
        # Clear previous figure state before plotting
        plt.cla()
        plt.clf()
        
        # Plot line chart
        plt.plot(time_data, value_data, label=f"Diameter: {diameter} m")
        plt.title("QGST Plot")
        plt.xlabel("Time (day)")
        plt.ylabel("QGST (sm³/d)")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.show()
        
        # Save chart
        plot_path = os.path.join(os.path.dirname(excel_path), "QGST_plot.png")
        plt.savefig(plot_path)
        plt.close()
        
        result["message"] = f"QGST chart generated: {plot_path}"
        result["plot_path"] = plot_path
        return result
    except Exception as e:
        result["status"] = "error"
        result["message"] = f"Failed to generate chart: {str(e)}"
        return result

@mcp.tool()
def generate_code(
    diameter: float = DEFAULT_DIAMETER, 
    pipeline_path: str = DEFAULT_MODEL_PATH,
    output_file: str = "olga_standalone.py",
    include_plot: bool = True
) -> dict:
    """
    Generate standalone executable code that performs OLGA simulation workflow
    Args:
        diameter: Pipeline diameter (unit: meter)
        pipeline_path: Directory path of the model files
        output_file: Output Python file name
        include_plot: Whether to include plot generation
    Returns:
        dict: Operation result with generated file path
    """
    result = {"status": "success", "message": ""}
    
    try:
        # Generate plot code conditionally
        plot_code = '''        # Step 4: Generate plot
        plot_path = generate_qgst_plot(DIAMETER, excel_path)''' if include_plot else '''        # Step 4: Plot generation disabled
        # plot_path = generate_qgst_plot(DIAMETER, excel_path)'''
        
        plot_print = '''        print(f"Plot saved to: {plot_path}")''' if include_plot else '''        # print(f"Plot saved to: {plot_path}")'''
        
        # Generate standalone code template
        code_template = f'''#!/usr/bin/env python3
"""
Standalone OLGA Simulation Script
Generated automatically - no MCP dependencies required
"""

import os
import matplotlib.pyplot as plt
import subprocess
from openpyxl import load_workbook
import OLGAOutput2xlsx

# Configuration
PIPELINE_PATH = r"{pipeline_path}"
DIAMETER = {diameter}
NTHREADS = {nthreads}
OLGA_BIN = r"{olga_bin}"
LICENSE_PATH = r"{LICENSE_PATH}"

def update_diameter(diameter, pipeline_path):
    """Update the pipeline diameter parameter in the Genkey file"""
    print(f"Updating diameter to {{diameter}} m...")
    
    genkey_file_path = os.path.join(pipeline_path, "Basic2.genkey")
    if not os.path.exists(genkey_file_path):
        raise FileNotFoundError(f"Genkey file does not exist: {{genkey_file_path}}")
    
    with open(genkey_file_path, "r", encoding="utf-8") as file:
        content = file.read()
    
    if "DIAMETER=" in content:
        updated_content = content.replace(
            f"DIAMETER={{content.split('DIAMETER=')[1].splitlines()[0]}}",
            f"DIAMETER={{diameter}} m"
        )
        with open(genkey_file_path, "w", encoding="utf-8") as file:
            file.write(updated_content)
        print(f"Successfully updated diameter to {{diameter}} m")
    else:
        raise ValueError("DIAMETER parameter not found in Genkey file")

def run_olga_simulation(pipeline_path):
    """Run OLGA simulation program"""
    print("Running OLGA simulation...")
    
    env = os.environ.copy()
    env["LM_LICENSE_FILE"] = LICENSE_PATH
    
    executed_files = []
    for subdir, _, files in os.walk(pipeline_path):
        for file in files:
            if file.endswith('.key') or file.endswith('.genkey'):
                full_path = os.path.join(subdir, file)
                cmd = f'"{{OLGA_BIN}}" -nthreads {{NTHREADS}} "{{full_path}}"'
                
                print(f"Executing: {{cmd}}")
                process = subprocess.run(
                    cmd,
                    env=env,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    check=True
                )
                executed_files.append(full_path)
    
    if not executed_files:
        print("Warning: No .key or .genkey files found")
    else:
        print(f"Successfully ran {{len(executed_files)}} simulation files")

def process_olga_results(pipeline_path):
    """Process OLGA output results and convert to Excel"""
    print("Processing OLGA results...")
    
    path_ = [pipeline_path]
    OLGAOutput2xlsx.testtpl(path_)
    
    excel_path = os.path.join(pipeline_path, "Basic2.tpl.xlsx")
    if os.path.exists(excel_path):
        print(f"Results converted to Excel: {{excel_path}}")
        return excel_path
    else:
        raise FileNotFoundError("Excel result file not generated")

def generate_qgst_plot(diameter, excel_path):
    """Generate QGST chart from Excel results"""
    print("Generating QGST plot...")
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file does not exist: {{excel_path}}")
    
    # Constant definitions
    SHEET_NAME = 'Data'
    TIME_COL = 2     # Column for time data (Column B)
    VALUE_COL = 23   # Column for value data (Column V)
    START_ROW = 8
    
    # Load workbook and worksheet
    wb = load_workbook(filename=excel_path)
    ws = wb[SHEET_NAME]
    
    # Read data
    time_data = []
    value_data = []
    
    for col in ws.iter_cols(min_row=START_ROW, min_col=TIME_COL, max_col=TIME_COL, values_only=True):
        time_data.extend([cell for cell in col if cell is not None])
    
    for col2 in ws.iter_cols(min_row=START_ROW, min_col=VALUE_COL, max_col=VALUE_COL, values_only=True):
        value_data.extend([cell for cell in col2 if cell is not None])
    
    # Ensure time data and value data have the same length
    min_len = min(len(time_data), len(value_data))
    time_data = time_data[:min_len]
    value_data = value_data[:min_len]
    
    # Clear previous figure state before plotting
    plt.cla()
    plt.clf()
    
    # Plot line chart
    plt.plot(time_data, value_data, label=f"Diameter: {{diameter}} m")
    plt.title("QGST Plot")
    plt.xlabel("Time (day)")
    plt.ylabel("QGST (sm³/d)")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    
    # Save chart
    plot_path = os.path.join(os.path.dirname(excel_path), "QGST_plot.png")
    plt.savefig(plot_path)
    plt.close()
    
    print(f"QGST chart generated: {{plot_path}}")
    return plot_path

def main():
    """Main execution function"""
    print("=" * 50)
    print("OLGA Simulation Workflow - Standalone Execution")
    print("=" * 50)
    
    try:
        # Step 1: Update diameter
        update_diameter(DIAMETER, PIPELINE_PATH)
        
        # Step 2: Run simulation
        run_olga_simulation(PIPELINE_PATH)
        
        # Step 3: Process results
        excel_path = process_olga_results(PIPELINE_PATH)
        
{plot_code}
        
        print("=" * 50)
        print("Workflow completed successfully!")
        print(f"Results saved to: {{excel_path}}")
{plot_print}
        print("=" * 50)
        
    except Exception as e:
        print(f"Error occurred: {{str(e)}}")
        raise

if __name__ == "__main__":
    main()
'''
        
        # Write the generated code to file
        output_path = os.path.join(os.getcwd(), output_file)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(code_template)
        
        # Make the file executable (Unix-like systems)
        if os.name != 'nt':  # Not Windows
            os.chmod(output_path, 0o755)
        
        result["message"] = f"Standalone code generated successfully: {output_path}"
        result["output_file"] = output_path
        result["parameters"] = {
            "diameter": diameter,
            "pipeline_path": pipeline_path,
            "include_plot": include_plot
        }
        
        return result
        
    except Exception as e:
        result["status"] = "error"
        result["message"] = f"Failed to generate code: {str(e)}"
        return result

# Start MCP Server
if __name__ == "__main__":
    print("Starting OLGA Backend Server...")
    mcp.run()
